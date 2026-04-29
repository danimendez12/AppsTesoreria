from datetime import datetime

import openpyxl
import models
import os
def ordenar_tabla(tabla):
    filas = list(tabla.iter_rows(min_row=2,values_only=True))
    filas.sort(key=lambda x: (x[2]))
    tabla.delete_rows(2, tabla.max_row)
    for fila in filas:
        tabla.append(fila)
    return tabla

def cargar_registros(archivo_excel):
    registros = []
    nombre = os.path.basename(archivo_excel)
    nombre = os.path.splitext(nombre)[0]
    df = openpyxl.load_workbook(archivo_excel)
    df1 = df.active
    sumatoriaMonto = 0
    sumatoriaCorrelativos = 0
    df1 = ordenar_tabla(df1)

    for row in df1.iter_rows(min_row=2, values_only=True):
        registro = models.Credito()
        carne = str(row[2])
        monto = float(row [7])
        banco = row [12]
        cuenta = str(row [13])
        if banco == "NO ASIGNADO" or monto <= 0:
            continue
        producto = cuenta[:3]
        moneda = cuenta[3:5]
        oficina = cuenta[5:8]
        numero_de_cuenta = cuenta[8:14]
        digito_verificador = cuenta[14]
        sumatoriaMonto += monto
        sumatoriaCorrelativos += int(numero_de_cuenta)
        registro.numero_de_cuenta = numero_de_cuenta
        registro.digito_verificador = digito_verificador
        registro.producto = producto
        registro.moneda = moneda
        registro.oficina_apertura = oficina
        registro.numero_comprobante = carne[2:10]
        monto_entero = int(round(monto * 100))
        registro.monto = str(monto_entero)
        registro.concepto_pago = f"CARNE: {carne}"
        registros.append(registro)
    registroDebito = models.Debito()
    registroDebito.oficina_apertura = "075"
    registroDebito.producto = "100"
    registroDebito.moneda = "01"
    registroDebito.numero_de_cuenta = "006662"
    registroDebito.digito_verificador = "1"
    registroDebito.numero_comprobante = "00000001"
    monto_float = float(sumatoriaMonto)
    monto_entero = int(round(sumatoriaMonto * 100))
    registroDebito.monto = str(monto_entero)
    fecha = datetime.now().strftime("%d/%m/%Y")
    registroDebito.concepto_pago = f"{nombre}{fecha}"
    registros.insert(0,registroDebito)
    sumatoriaMonto += monto_float
    sumatoriaCorrelativos += int("006662")


    return registros, sumatoriaMonto, sumatoriaCorrelativos

def generar_encabezado():
    encabezado = models.Encabezado(
        tipo="1",
        num_cliente="023232",
        fecha= datetime.now().strftime("%d%m%Y"),
        num_transferencia_real= "000000",
        num_transferencia_interna= "000000",
        tipo_transaccion= "1",
        codigo_error= "0000",
        total_transferencia= "000000000000",
        tipo_cambio_compra_dia= "000000",
        tipo_cambio_venta_dia= "0000000",
        campo_sin_uso= ""
    )
    return encabezado

def generar_registro_control(sumatoriaMonto, sumatoriaCorrelativos):
    registro_control = models.RegistroControl(
        tipo="4",
        sumatoria_montos= str(int(round(sumatoriaMonto * 100))),
        sumatoria_correlativos= str(sumatoriaCorrelativos),
        campo_sin_uso= "",
        test_key= "",
        monto_en_dolares= "",
        monto_en_colones= "",
    )
    return registro_control

def generar_txt(encabezado,registros,registro_control):
    transaccion = models.Transaccion(
        encabezado=encabezado,
        registros=registros,
        registroControl=registro_control
    )
    return transaccion.to_env()